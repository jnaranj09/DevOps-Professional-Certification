from openpyxl import Workbook, load_workbook

# Cargamos el archivo excel
filesheet = "Workspace\Archivos-excel\DB.xlsx"
wb = load_workbook(filesheet)
ws = wb.active

# ws['A1'] = "Cedula"
# ws['B1'] = "Nombre"
# ws['C1'] = "Promedio Final"
# ws['D1'] = "Estado"

# get current row
row = ws.max_row

ws.append(
    ["=RANDBETWEEN(111111111,777777777)", "Juan", "=RANDBETWEEN(0,100)", '=IF(' + 'C' + str(row+1) + '>=70,"Aprobado","Reprobado")'],
)

wb.save(filesheet)

print (ws['A'+ str(row+1)].value + " | " + ws['B'+ str(row+1)].value + " | " + ws['C'+ str(row+1)].value + " | " + ws['D'+ str(row+1)].value)
